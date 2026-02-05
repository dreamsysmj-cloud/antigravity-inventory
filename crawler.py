import time
import os
import glob
import shutil
import pandas as pd
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import NoSuchElementException, NoAlertPresentException, UnexpectedAlertPresentException, StaleElementReferenceException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl.styles import Border, Side

# ======================================================
# 👇 비밀번호는 맨 아래쪽에서 입력해주세요!
# ======================================================

BASE_DIR = os.getcwd()
DOWNLOAD_DIR = os.path.join(BASE_DIR, "data")
if not os.path.exists(DOWNLOAD_DIR):
    os.makedirs(DOWNLOAD_DIR)

# ------------------------------------------------------
# [핵심] 브라우저 설정
# ------------------------------------------------------
def get_fresh_driver():
    options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": DOWNLOAD_DIR,
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False,
        "profile.password_manager_leak_detection": False,
        "safebrowsing.enabled": True
    }
    options.add_experimental_option("prefs", prefs)
    options.add_argument("--disable-save-password-bubble")
    options.add_argument("--disable-features=PasswordLeakDetection")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_argument("--disable-blink-features=AutomationControlled")
    
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.maximize_window()
    return driver

# ------------------------------------------------------
# [공통] 안전 클릭
# ------------------------------------------------------
def safe_click_text(driver, text_list):
    try: webdriver.ActionChains(driver).send_keys(Keys.ESCAPE).perform()
    except: pass
    time.sleep(0.5)
    
    for text in text_list:
        try:
            xpath = f"//*[contains(text(), '{text}')]"
            elements = driver.find_elements(By.XPATH, xpath)
            for element in elements:
                if element.is_displayed():
                    driver.execute_script("arguments[0].click();", element)
                    print(f"      👉 '{text}' 클릭 성공")
                    time.sleep(2) 
                    return True
        except:
            continue
    return False

# ------------------------------------------------------
# [공통] 파일 다운로드 및 이름 변경
# ------------------------------------------------------
def wait_for_new_file_and_rename(company_name, before_files, suffix="재고"):
    print(f"   ⏳ [{company_name} {suffix}] 다운로드 확인 중...")
    
    target_file = None
    for i in range(30):
        time.sleep(1)
        current_files = glob.glob(os.path.join(DOWNLOAD_DIR, "*"))
        current_files = [f for f in current_files if f.endswith(".xlsx") or f.endswith(".xls")]
        
        new_files = list(set(current_files) - set(before_files))
        
        if new_files:
            if any("crdownload" in f or "tmp" in f for f in new_files):
                continue
            target_file = new_files[0]
            print(f"   ✅ 파일 다운로드 완료: {os.path.basename(target_file)}")
            break
    
    if target_file:
        time.sleep(2)
        fixed_name = f"latest_{company_name}{suffix}.xlsx"
        fixed_path = os.path.join(DOWNLOAD_DIR, fixed_name)

        try:
            if os.path.exists(fixed_path):
                os.remove(fixed_path)
            os.rename(target_file, fixed_path)
            print(f"   🎉 저장 완료: {fixed_name}")
        except: pass
    else:
        print(f"   ❌ [오류] {company_name} {suffix}: 파일을 못 받았습니다 (시간 초과).")

# ------------------------------------------------------
# [프로세스] 이카운트 (재고 + 판매)
# ------------------------------------------------------
def run_ecount_task(com_code, user_id, user_pw, company_name):
    print(f"\n🚀 [{company_name}] 시작...")
    driver = get_fresh_driver()
    
    # [1] 로그인
    try:
        driver.get("https://login.ecount.com/")
        driver.implicitly_wait(5)
        
        driver.find_element(By.ID, "com_code").clear()
        driver.find_element(By.ID, "com_code").send_keys(com_code)
        driver.find_element(By.ID, "id").clear()
        driver.find_element(By.ID, "id").send_keys(user_id)
        driver.find_element(By.ID, "passwd").clear()
        driver.find_element(By.ID, "passwd").send_keys(user_pw)
        driver.find_element(By.ID, "passwd").send_keys(Keys.ENTER)
        
        try: 
            WebDriverWait(driver, 3).until(EC.alert_is_present())
            driver.switch_to.alert.accept()
        except: pass
        time.sleep(3) 

        # -------------------------------------------------
        # [Step 1] 재고현황 다운로드
        # -------------------------------------------------
        files_before = glob.glob(os.path.join(DOWNLOAD_DIR, "*"))
        files_before = [f for f in files_before if f.endswith(".xlsx") or f.endswith(".xls")]

        if company_name == "하은재고":
            safe_click_text(driver, ["MyPage", "Mypage"])
            safe_click_text(driver, ["재고현황"])
        elif company_name == "한국재고":
            webdriver.ActionChains(driver).send_keys(Keys.ESCAPE).perform()
            safe_click_text(driver, ["재고I", "재고 I", "재고1", "재고 1"])
            safe_click_text(driver, ["재고현황"])
        elif company_name == "다이소재고":
            safe_click_text(driver, ["재고I", "재고 I", "재고1", "재고 1"])
            time.sleep(1)
            safe_click_text(driver, ["출력물", "출 력 물"])

        time.sleep(2)
        # 검색(F8)
        if not safe_click_text(driver, ["검색(F8)", "검색", "F8"]):
             webdriver.ActionChains(driver).send_keys(Keys.F8).perform()
        
        time.sleep(3)
        # 엑셀 다운
        safe_click_text(driver, ["Excel", "EXCEL", "엑셀"])
        
        wait_for_new_file_and_rename(company_name, files_before, suffix="재고")

        # -------------------------------------------------
        # [Step 2] 판매현황 다운로드
        # -------------------------------------------------
        print(f"      🕵️‍♀️ [{company_name}] '판매현황' 이동 중...")
        files_before_sales = glob.glob(os.path.join(DOWNLOAD_DIR, "*"))
        files_before_sales = [f for f in files_before_sales if f.endswith(".xlsx") or f.endswith(".xls")]
        
        # 메뉴 찾기 (ESC로 팝업 닫고 시작)
        webdriver.ActionChains(driver).send_keys(Keys.ESCAPE).perform()
        time.sleep(1)

        # 판매현황 버튼 클릭
        if safe_click_text(driver, ["판매현황", "판매 현황"]):
            time.sleep(2)
            
            # [NEW] 날짜 설정: "금월(~오늘)" 버튼 클릭
            # '금월' 또는 '금월(~오늘)' 텍스트가 포함된 버튼/링크 클릭
            if safe_click_text(driver, ["금월(~오늘)", "금월", "Today", "This Month"]):
                 print("      📅 날짜를 '금월(This Month)'로 설정했습니다.")
                 time.sleep(2)
            else:
                 print("      ⚠️ '금월' 설정 버튼을 찾지 못해 기본 날짜로 진행합니다.")

            # 검색(F8)
            if not safe_click_text(driver, ["검색(F8)", "검색", "F8"]):
                 webdriver.ActionChains(driver).send_keys(Keys.F8).perform()
            
            time.sleep(3)
            # EXCEL(화면) 클릭
            if not safe_click_text(driver, ["EXCEL(화면)", "Excel(화면)", "엑셀(화면)"]):
                safe_click_text(driver, ["Excel", "EXCEL", "엑셀"])
            
            # 다운로드 (필터링 없이 이름만 변경)
            wait_for_new_file_and_rename(company_name, files_before_sales, suffix="판매")
            
        else:
            print("      ⚠️ '판매현황' 메뉴를 찾지 못했습니다.")

    except Exception as e:
        print(f"   ❌ [{company_name}] 오류: {e}")
    finally:
        driver.quit()

# ------------------------------------------------------
# [NEW] 파일 통합 (재고 + 판매)
# ------------------------------------------------------
# ------------------------------------------------------
# [NEW] 파일 통합 (재고 + 판매)
# ------------------------------------------------------
def merge_all_files():
    print("\n📚 [마무리] 파일 통합 중...")
    
    now_dt = datetime.now()
    # 포맷: 260204-18시30분 통합데이터.xlsx
    filename_str = now_dt.strftime("%y%m%d-%H시%M분")
    target_filename = f"{filename_str} 통합데이터.xlsx"
    
    # -------------------------------------------------
    # [Start] 폴더 구조 생성 로직
    # data -> 26년2월 -> 2월4일 -> 파일
    # -------------------------------------------------
    year_suffix = now_dt.strftime("%y") # 26
    month = now_dt.month # 2
    day = now_dt.day # 4
    
    folder_year_month = f"{year_suffix}년{month}월"
    folder_day = f"{month}월{day}일"
    
    # 최종 저장 경로: data/26년2월/2월4일/
    save_dir = os.path.join(DOWNLOAD_DIR, folder_year_month, folder_day)
    
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
        print(f"   📂 폴더 생성: {save_dir}")
        
    target_file = os.path.join(save_dir, target_filename)
    # -------------------------------------------------
    
    # (시트명, 파일명) 매핑 - latest 파일은 data 폴더(=DOWNLOAD_DIR)에 그대로 있음
    files_map = {
        "하은재고": "latest_하은재고재고.xlsx",
        "한국재고": "latest_한국재고재고.xlsx",
        "다이소재고": "latest_다이소재고재고.xlsx",
        "하은판매": "latest_하은재고판매.xlsx",
        "한국판매": "latest_한국재고판매.xlsx",
        "다이소판매": "latest_다이소재고판매.xlsx",
    }
    
    try:
        with pd.ExcelWriter(target_file, engine='openpyxl') as writer:
            merged_count = 0
            for sheet_name, filename in files_map.items():
                # latest 파일은 원래 위치(DOWNLOAD_DIR)에서 찾음
                file_path = os.path.join(DOWNLOAD_DIR, filename)
                if os.path.exists(file_path):
                    try:
                        df = pd.read_excel(file_path, header=None) # 헤더 없이 읽어서 그대로 붙여넣기
                        
                        df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                        
                        # [NEW] 테두리 추가 로직
                        ws = writer.sheets[sheet_name]
                        border_style = Border(
                            left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin')
                        )
                        for row in ws.iter_rows():
                            for cell in row:
                                cell.border = border_style

                        # [NEW] 열 너비 자동 조정 (A열 제외)
                        for col in ws.columns:
                            try:
                                column_letter = col[0].column_letter # A, B, C...
                            except:
                                # column_letter 속성이 없는 구버전 대비
                                from openpyxl.utils import get_column_letter
                                column_letter = get_column_letter(col[0].column)

                            if column_letter == 'A':
                                continue
                                
                            max_length = 0
                            for cell in col:
                                try:
                                    if cell.value:
                                        # 한글 길이 보정을 위해 문자열 순회
                                        cell_len = 0
                                        for char in str(cell.value):
                                            if ord(char) > 127: # 한글 등 멀티바이트 문자
                                                cell_len += 1.7 
                                            else:
                                                cell_len += 1.0
                                        
                                        if cell_len > max_length:
                                            max_length = cell_len
                                except: pass
                            
                            # 여유 공간 추가
                            adjusted_width = max_length + 2
                            # 최소 너비 설정
                            if adjusted_width < 10: adjusted_width = 10
                            
                            ws.column_dimensions[column_letter].width = adjusted_width

                        print(f"   ✅ 시트 추가: {sheet_name}")
                        merged_count += 1
                    except: pass
            
        if merged_count > 0:
            print(f"🎉 통합 완료! -> {target_file}")
            
    except Exception as e:
        print(f"   ❌ 통합 저장 실패: {e}")

# ======================================================
# 🏁 실행
# ======================================================
def run_independent_mode():
    print("🤖 통합 로봇 가동 (이카운트 3사: 재고 + 판매)")

    # 1. 하은
    run_ecount_task("83666", "한국사료", PASSWORD_HAEUN, "하은재고")
    time.sleep(1)

    # 2. 한국
    run_ecount_task("89587", "or7lsarang", PASSWORD_HANKOOK, "한국재고")
    time.sleep(1)

    # 3. 다이소
    run_ecount_task("69903", "한국사료", PASSWORD_DAISO, "다이소재고")
    time.sleep(1)

    # 4. 통합
    merge_all_files()

    print("\n👋 모든 작업 종료!")

# ======================================================
# 👇 비밀번호 입력 (여기에 입력하세요)
# ======================================================
PASSWORD_HAEUN = "cccc1004"
PASSWORD_HANKOOK = "thedog1044!!"
PASSWORD_DAISO = "gksrnrtkfy1234"

if __name__ == "__main__":
    run_independent_mode()
